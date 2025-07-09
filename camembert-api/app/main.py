from fastapi import FastAPI, HTTPException, UploadFile, File,BackgroundTasks
from openpyxl import load_workbook
import io
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
import torch
from transformers import CamembertTokenizer, CamembertModel
import torch.nn as nn
import uvicorn
from typing import Dict
import logging
import os
from .config import settings
import requests

# Configure logging
logging.basicConfig(level=getattr(logging, settings.LOG_LEVEL))
logger = logging.getLogger(__name__)

# Initialize FastAPI app
app = FastAPI(
    title="CamemBERT Multi-Output Prediction API",
    description="API for predicting Fiabilité Intégrité, Disponibilité, and Process Safety scores",
    version="1.0.0",
    docs_url="/docs",
    redoc_url="/redoc"
)

# Add CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=settings.ALLOWED_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Device setup
device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
logger.info(f"Using device: {device}")

# Define the model structure
class CamemBERTMultiOutput(nn.Module):
    def __init__(self):
        super().__init__()
        self.camembert = CamembertModel.from_pretrained("camembert-base")
        self.dropout = nn.Dropout(0.3)
        self.heads = nn.ModuleList([
            nn.Linear(self.camembert.config.hidden_size, 6) for _ in range(3)
        ])  # 3 tasks, 6 classes each (0-5)

    def forward(self, input_ids, attention_mask):
        output = self.camembert(input_ids=input_ids, attention_mask=attention_mask)
        cls_output = self.dropout(output.last_hidden_state[:, 0, :])
        logits = [head(cls_output) for head in self.heads]
        return logits

# Global variables for model and tokenizer
model = None
tokenizer = None

# Pydantic models for request/response
class PredictionRequest(BaseModel):
    description_anomaly: str
    description_equipement: str

class PredictionResponse(BaseModel):
    fiabilite_integrite: int
    disponibilite: int
    process_safety: int
    criticite: int

@app.on_event("startup")
async def load_model():
    """Load the model and tokenizer on startup"""
    global model, tokenizer
    
    try:
        logger.info("Loading tokenizer...")
        tokenizer = CamembertTokenizer.from_pretrained("camembert-base")
        
        logger.info(f"Loading model from {settings.MODEL_PATH}...")
        model = CamemBERTMultiOutput().to(device)
        
        # Check if model file exists
        if not os.path.exists(settings.MODEL_PATH):
            raise FileNotFoundError(f"Model file not found: {settings.MODEL_PATH}")
            
        model.load_state_dict(torch.load(settings.MODEL_PATH, map_location=device))
        model.eval()
        
        logger.info("Model and tokenizer loaded successfully!")
        
    except Exception as e:
        logger.error(f"Error loading model: {str(e)}")
        raise e

def clean_text(text):
    text = str(text).lower()
    text = text.replace('\n', ' ').replace('\t', ' ')
    return ' '.join(text.split())

def predict_text(text: str) -> Dict[str, int]:
    """Make prediction on input text"""
    text = clean_text(text)
    if model is None or tokenizer is None:
        raise HTTPException(status_code=500, detail="Model not loaded")
    
    try:
        # Tokenize input
        encoding = tokenizer(
            text, 
            return_tensors="pt", 
            max_length=settings.MAX_LENGTH,
            padding="max_length", 
            truncation=True
        )
        
        input_ids = encoding["input_ids"].to(device)
        attention_mask = encoding["attention_mask"].to(device)

        # Make prediction
        with torch.no_grad():
            outputs = model(input_ids, attention_mask)

        # Get predictions
        predictions = [torch.argmax(logit, dim=1).item() for logit in outputs]
        
        return {
            "fiabilite_integrite": predictions[0],
            "disponibilite": predictions[1],
            "process_safety": predictions[2],
            "criticite": predictions[0] + predictions[1] + predictions[2]
        }
        
    except Exception as e:
        logger.error(f"Error during prediction: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Prediction failed: {str(e)}")

@app.get("/")
async def root():
    """Root endpoint with API information"""
    return {
        "message": "CamemBERT Multi-Output Prediction API",
        "status": "running",
        "device": str(device),
        "model_loaded": model is not None,
        "version": "1.0.0"
    }

@app.get("/health")
async def health_check():
    """Health check endpoint"""
    return {
        "status": "healthy",
        "model_loaded": model is not None,
        "tokenizer_loaded": tokenizer is not None,
        "device": str(device)
    }

@app.post("/predict", response_model=PredictionResponse)
async def predict(request: PredictionRequest):
    """
    Predict scores for the given text
    
    - **text**: Input text to analyze
    
    Returns predictions for:
    - **fiabilite_integrite**: Score from 0-5
    - **disponibilite**: Score from 0-5  
    - **process_safety**: Score from 0-5
    - **criticite**: Sum of all three scores
    """
    if not request.description_equipement.strip() or not request.description_anomaly.strip():
        raise HTTPException(status_code=400, detail="Text cannot be empty")
    
    try:
        # Combine 'Description de l\'équipement' and 'Description' into a single text
        combined_text = f"{request.description_equipement} {request.description_anomaly}"
        predictions = predict_text(combined_text)
        return PredictionResponse(**predictions)
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Unexpected error: {str(e)}")
        raise HTTPException(status_code=500, detail="Internal server error")

@app.post("/batch_predict")
async def batch_predict(texts: list[str]):
    """
    Predict scores for multiple texts
    
    - **texts**: List of input texts to analyze
    
    Returns a list of predictions for each text
    """
    if not texts:
        raise HTTPException(status_code=400, detail="Text list cannot be empty")
    
    if len(texts) > settings.BATCH_SIZE_LIMIT:
        raise HTTPException(status_code=400, detail=f"Batch size too large (max {settings.BATCH_SIZE_LIMIT})")
    
    try:
        results = []
        for text in texts:
            if not text.strip():
                continue
            predictions = predict_text(text)
            results.append({
                "text": text,
                "predictions": predictions
            })
        
        return {"results": results}
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Unexpected error in batch prediction: {str(e)}")
        raise HTTPException(status_code=500, detail="Internal server error")


# Background processing function
def process_excel_file(file_bytes: bytes):
    try:
        workbook = load_workbook(io.BytesIO(file_bytes))
        sheet = workbook.active

        headers = [cell.value for cell in sheet[1]]
        required_columns = [
            "Num_equipement",
            "Systeme", 
            "Description",
            "Date de détéction de l'anomalie",
            "Description de l'équipement",
            "Section propriétaire"
        ]

        column_indices = {}
        for col in required_columns:
            if col not in headers:
                logger.warning(f"Missing required column: {col}")
                return
            column_indices[col] = headers.index(col)

        results = []

        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            description = row[column_indices["Description"]]
            
            if description:
                predictions = predict_text(description)
                row_data = {
                'num_equipments': str(row[column_indices["Num_equipement"]]),
                'systeme': str(row[column_indices["Systeme"]]),
                'descreption_anomalie': str(row[column_indices["Description"]]),
                'date_detection': str(row[column_indices["Date de détéction de l'anomalie"]]),
                'descreption_equipment': str(row[column_indices["Description de l'équipement"]]),
                'section_proprietaire': str(row[column_indices["Section propriétaire"]]),
                'fiablite_integrite': str(predictions['fiabilite_integrite']),
                'disponsibilite': str(predictions['disponibilite']),
                'process_safty': str(predictions['process_safety']),
                'criticite': str(predictions['criticite'])
                }
                results.append(row_data)

            # Send results to external API
            if row_idx % 5 == 0:
                try:
                    response = requests.post("http://backend:7532/ml/ml-response", json=results)
                    print(results)
                    response.raise_for_status()
                    logger.info(f"Posted results to external API: {response.status_code}")
                    results.clear()  # Clear results after posting
                    # break
                except requests.RequestException as e:
                    logger.error(f"Failed to post results to external API: {e}")
                    # break
            if row_idx % 50 == 0:
                break


        # TODO: save results to DB, file, email, etc.
        logger.info(f"Processed {len(results)} rows in background.")
        if (len(results)):
            logger.info(f"Processed {results[0]}")

    except Exception as e:
        logger.error(f"Error in background task: {e}")

# Endpoint that triggers background task
@app.post("/predict_excel")
async def predict_excel(background_tasks: BackgroundTasks, file: UploadFile = File(...)):
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="File must be Excel format (.xlsx or .xls)")

    contents = await file.read()
    background_tasks.add_task(process_excel_file, contents)

    return {
        "message": "File received. Processing will continue in background."
    }

if __name__ == "__main__":
    uvicorn.run(
        "app.main:app",
        host=settings.HOST,
        port=settings.PORT,
        reload=False
    )